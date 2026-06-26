"""文件解析器 — 支持 Excel (.xlsx/.xls) 和 CSV 文件"""

import csv
import re
import pandas as pd
from pathlib import Path


def parse_file(file_path: str) -> tuple[list[str], list[list]]:
    """
    解析上传文件，自动识别表头行并返回数据。

    Args:
        file_path: 文件路径

    Returns:
        (headers: list[str], rows: list[list]) — 表头列表 + 数据行列表
    """
    ext = Path(file_path).suffix.lower()

    if ext == ".csv":
        return _parse_csv(file_path)
    elif ext in (".xlsx", ".xls"):
        return _parse_excel(file_path)
    else:
        raise ValueError(f"不支持的文件格式: {ext}（仅支持 .xlsx .xls .csv）")


def _parse_csv(file_path: str) -> tuple[list[str], list[list]]:
    """解析 CSV 文件"""
    # 先探测编码
    encodings = ["utf-8", "utf-8-sig", "gbk", "gb2312", "gb18030", "latin-1"]
    content = None
    used_encoding = "utf-8"

    for enc in encodings:
        try:
            with open(file_path, "r", encoding=enc, newline="") as f:
                content = f.read()
            used_encoding = enc
            break
        except (UnicodeDecodeError, UnicodeError):
            continue

    if content is None:
        raise ValueError("无法识别 CSV 文件编码")

    lines = content.strip().splitlines()
    reader = csv.reader(lines)
    all_rows = list(reader)

    if len(all_rows) < 2:
        raise ValueError("CSV 文件至少需要表头行 + 一行数据")

    header_idx = _detect_header_row(all_rows)
    headers = [_clean_header(h) for h in all_rows[header_idx]]
    data_rows = all_rows[header_idx + 1 :]

    # 清理：移除全空行、移除空值填充
    data_rows = [
        [cell.strip() if isinstance(cell, str) else cell for cell in row]
        for row in data_rows
        if any(
            cell is not None and str(cell).strip() != "" for cell in row
        )
    ]

    return headers, data_rows


def _clean_header(header: str | None) -> str:
    """清洗表头：去 BOM、去空格、去引号"""
    if header is None:
        return ""
    s = str(header).strip()
    # 移除 UTF-8 BOM
    if s.startswith("\ufeff"):
        s = s[1:]
    # 移除零宽字符
    s = s.replace("\u200b", "").replace("\ufeff", "")
    # 移除首尾引号
    s = s.strip('"').strip("'").strip()
    return s


def _parse_excel(file_path: str) -> tuple[list[str], list[list]]:
    """解析 Excel 文件"""
    ext = Path(file_path).suffix.lower()
    # .xls 旧格式用 xlrd，.xlsx/.xlsm 用 openpyxl
    if ext == ".xls" and _is_old_xls(file_path):
        try:
            return _parse_excel_xlrd(file_path)
        except Exception:
            return _parse_excel_pandas(file_path)
    # .xlsx/.xlsm 或无法判断：先用 openpyxl，pandas 做 fallback
    try:
        return _parse_excel_openpyxl(file_path)
    except Exception:
        return _parse_excel_pandas(file_path)


def _parse_excel_openpyxl(file_path: str) -> tuple[list[str], list[list]]:
    """用 openpyxl 解析（更精确控制表头识别）"""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))

    wb.close()

    if len(all_rows) < 2:
        raise ValueError("Excel 文件至少需要表头行 + 一行数据")

    header_idx = _detect_header_row(all_rows)
    headers = [_clean_header(h) for h in all_rows[header_idx]]
    data_rows = all_rows[header_idx + 1 :]

    # 清理：移除全空行
    data_rows = [
        [cell if cell is not None else "" for cell in row]
        for row in data_rows
        if any(cell is not None and str(cell).strip() != "" for cell in row)
    ]

    return headers, data_rows


def _parse_excel_pandas(file_path: str) -> tuple[list[str], list[list]]:
    """用 pandas 解析（fallback）"""
    df = pd.read_excel(file_path, header=None, dtype=str)
    all_rows = df.fillna("").values.tolist()

    if len(all_rows) < 2:
        raise ValueError("Excel 文件至少需要表头行 + 一行数据")

    header_idx = _detect_header_row(all_rows)
    headers = [_clean_header(h) for h in all_rows[header_idx]]
    data_rows = all_rows[header_idx + 1 :]

    data_rows = [
        [str(cell).strip() for cell in row]
        for row in data_rows
        if any(str(cell).strip() != "" for cell in row)
    ]

    return headers, data_rows


# ── TASK-078：多行复合表头识别 ──────────────────────────
# 科目余额表常见多行表头：
#   row 0: 科目类别 | 科目编码 | 科目名称 | 币种 | 期初余额 | 本期发生 | ...
#   row 1:                                 借方    贷方
#   row 2:                                 数量    金额
#   row 3+: 数据
# 本组工具用于：识别表头说明行、计算 data_start_row、合并出可读列名。


# 命中即判定为「表头单元格」的关键词（按出现字符串）。
_HEADER_KEYWORDS = (
    "科目", "余额", "金额", "借方", "贷方", "发生", "累计",
    "期初", "期末", "年初", "年末", "数量", "币种", "公司",
    "核算项目", "年", "级次",
)

# 强信号：出现以下关键词的行可认定为主表头行（携带「科目xx」列名 + 期间大类）。
_HEADER_ANCHOR_KEYWORDS = (
    "科目编码", "科目代码", "科目编号", "科目号", "科目名称",
    "科目名", "科目全称", "科目类别",
)

# 明显数据片段：纯数字、长客户名称特征、"茂名市" 等公司名后缀。行只要含其中一项，
# 即视为数据行（而非表头子行）。
_DATA_COMPANY_SUFFIX_MARKERS = ("公司", "工厂", "银行", "支行", "营业部", "基金", "信托")


def _row_has_header_anchor(row: list) -> bool:
    """该行是否含科目编码/科目名称等列名（主表头行判定）。"""
    for cell in row:
        if cell is None:
            continue
        s = str(cell).strip()
        if not s or s == "None":
            continue
        if any(k in s for k in _HEADER_ANCHOR_KEYWORDS):
            return True
    return False


def _header_cell_score(cell) -> int:
    """单元格内容是否像表头文本，返回分数（0/4/5）。"""
    if cell is None:
        return 0
    s = str(cell).strip()
    if not s or s == "None":
        return 0
    # 含「科目xx」的列名最强信号
    if any(k in s for k in _HEADER_ANCHOR_KEYWORDS):
        return 5
    if any(k in s for k in _HEADER_KEYWORDS):
        return 4
    return 0


def _row_total_header_score(row: list) -> int:
    """整行表头分：所有单元格 _header_cell_score 之和。"""
    return sum(_header_cell_score(c) for c in row)


def _row_is_data(row: list) -> bool:
    """判断一行是否像数据行（含科目代码数字、辅助明细方括号、或公司名后缀）。"""
    for cell in row:
        s = "" if cell is None else str(cell).strip()
        if not s or s == "None":
            continue
        # 全/半角空白开头 + 长公司名 → 辅助核算数据行
        if s.startswith(("\u3000", " ")) and ("[" in s or any(m in s for m in _DATA_COMPANY_SUFFIX_MARKERS)):
            return True
        # 开头带方括号 [辅助编码]
        if s.startswith("["):
            return True
        # 纯科目代码（数字 / 带分隔符的小数点代码）
        if re.fullmatch(r"\d+(?:[.\-]?\d+)*", s):
            return True
        # 开头是全角空白 + 数字（广西的" 1002001"）
        if (
            s.startswith("\u3000")
            and re.fullmatch(r"\u3000+\d+(?:[.\-]?\d+)*", s)
        ):
            return True
    return False


def detect_header_area(rows: list[list], max_scan: int = 15) -> tuple[int, int, list[int]]:
    """识别表头区域。

    返回 (main_header_row, data_start_row, header_rows)：
      - main_header_row: 主表头行（最优含「科目编码/科目名称」锚点的行）
      - data_start_row: 第一个数据行
      - header_rows: 主表头及紧接其后的表头说明子行（如借/贷、数量/金额）
    """
    scan_end = min(max_scan, len(rows))

    # 主表头行优先选取含「科目编码/科目名称」锚点的行（再按分值最高的）
    anchor_idxs = [i for i in range(scan_end) if _row_has_header_anchor(rows[i])]
    if anchor_idxs:
        main = max(anchor_idxs, key=lambda i: _row_total_header_score(rows[i]))
    else:
        # 无锚点：按总分最大的行
        main = max(range(scan_end), key=lambda i: _row_total_header_score(rows[i]))

    header_rows = [main]
    j = main + 1
    while j < len(rows):
        row = rows[j]
        if all(c is None or str(c).strip() == "" for c in row):
            j += 1
            continue
        if _row_is_data(row):
            break
        header_rows.append(j)
        j += 1
    data_start_row = j
    return main, data_start_row, header_rows


def _clean_for_header(value) -> str:
    """表头单元格清洗：None→''，去 BOM/零宽/空白。"""
    if value is None:
        return ""
    s = str(value).replace("\ufeff", "").replace("\u200b", "").strip()
    if s == "None":
        return ""
    return s


def merge_multirow_headers(all_rows: list[list], header_rows: list[int]) -> list[str]:
    """把多行表头合并成单行可读列名。

    规则：
    - 主表头行（header_rows[0]）空白单元格用左侧最近非空值回填（处理跨列合并标题，
      如 "期初余额" 跨 4 列）。
    - 子表头行（借/贷、数量/金额）按列追加去重；同时每个子行也做左回填，
      保证 "借方" 子标签覆盖相邻的 "数量/金额" 单元格。
    - 例：广西 col idx6 → "期初余额_借方_金额"；金蝶 col idx8 → "期初余额_借方"。
    """
    if not header_rows or not all_rows:
        return ["" for _ in (all_rows[0] if all_rows else [])]
    ncols = max((len(all_rows[i]) for i in header_rows), default=0)

    def _left_fill(row: list) -> list[str]:
        out: list[str] = []
        fillval = ""
        for j in range(ncols):
            v = _clean_for_header(row[j]) if j < len(row) else ""
            # 下一区块需重置：遇到表头大类关键字（如「期初余额」「本期发生」）
            # 出现在另一行同列时不切；这里仅做单行内左回填即可。
            if v:
                fillval = v
            out.append(fillval)
        return out

    def _left_fill_main(row: list) -> list[str]:
        """主行回填：在大类关键字切换时重置。

        例：主文「期初余额」跨 4 列应回填为 4 个「期初余额」；
        到了「本期发生」列自然替换。
        """
        out: list[str] = []
        fillval = ""
        for j in range(ncols):
            v = _clean_for_header(row[j]) if j < len(row) else ""
            if v:
                fillval = v
            out.append(fillval)
        return out

    filled_main = _left_fill_main(all_rows[header_rows[0]])
    filled_children = [_left_fill(all_rows[ri]) for ri in header_rows[1:]]

    merged: list[str] = []
    for j in range(ncols):
        parts: list[str] = []
        if filled_main[j]:
            parts.append(filled_main[j])
        for fc in filled_children:
            v = fc[j] if j < len(fc) else ""
            if v and v not in parts:
                parts.append(v)
        merged.append("_".join(parts))
    return merged


def _detect_header_row(rows: list[list], max_scan: int = 5) -> int:
    """
    自动检测表头行：扫描前 N 行，找到第一个所有单元格都非空的行。

    规则：
    1. 扫描前 max_scan 行
    2. 找到非空率最高的行
    3. 如果第一行非空率 ≥ 50%，默认第一行就是表头
    """
    best_idx = 0
    best_count = 0

    scan_end = min(max_scan, len(rows))
    for i in range(scan_end):
        row = rows[i]
        non_empty = sum(1 for cell in row if cell is not None and str(cell).strip() != "")
        if non_empty > best_count:
            best_count = non_empty
            best_idx = i

    # 优先第一行（大多数财务软件第一行就是表头）
    if best_idx == 0 and best_count > 0:
        return 0

    # 如果找到的行非空数明显更多，用那一行
    first_row_non_empty = sum(
        1 for cell in rows[0] if cell is not None and str(cell).strip() != ""
    )
    if best_count > first_row_non_empty * 1.5:
        return best_idx

    return 0


def parse_file_info(file_path: str) -> dict:
    """获取文件基本信息（用于预览前确认）"""
    headers, rows = parse_file(file_path)
    return {
        "file_name": Path(file_path).name,
        "headers": headers,
        "header_count": len(headers),
        "row_count": len(rows),
        "preview_rows": rows[:5],
    }


# ── TASK-078：科目余额表标准化导入专用解析 ────────────────


def parse_trial_balance_import(file_path: str) -> dict:
    """读取科目余额表原始行，识别多行表头区域，返回合并后的可读列名 + 数据起始行。

    返回：
        {
            "all_rows": list[list],          # 工作表全部行
            "merged_headers": list[str],     # 合并后可读列名
            "header_rows": list[int],        # 表头说明行（工作表行索引）
            "data_start_row": int,           # 第一个数据行（工作表行索引）
        }
    """
    ext = Path(file_path).suffix.lower()
    if ext in (".xlsx", ".xls"):
        _, all_rows = _parse_excel_all_rows(file_path)
    elif ext == ".csv":
        _, all_rows = _parse_csv_all_rows(file_path, "auto")
    else:
        raise ValueError(f"不支持的文件格式: {ext}")

    if len(all_rows) < 2:
        raise ValueError("Excel 文件至少需要表头行 + 一行数据")

    main, data_start, header_rows = detect_header_area(all_rows)
    merged = merge_multirow_headers(all_rows, header_rows)
    return {
        "all_rows": all_rows,
        "merged_headers": merged,
        "header_rows": header_rows,
        "data_start_row": data_start,
    }


def slice_data_rows(
    all_rows: list[list],
    data_start_row: int,
) -> list[list]:
    """从工作表行集合切片出数据行（与解析逻辑一致的清洗：空 None→''，跳过全空行）。

    返回行的顺序保持，索引即 data row 的 0 起序号。
    """
    section = all_rows[max(data_start_row, 0):]
    out: list[list] = []
    for row in section:
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        out.append(["" if c is None else c for c in row])
    return out


def parse_file_with_config(
    file_path: str,
    parse_config: dict | None = None,
) -> tuple[list[str], list[list]]:
    """
    按模板 parse_config 解析文件。

    parse_config 支持的字段：
    - header_row: int (0 基) — 表头行，默认 0
    - data_start_row: int (0 基) — 数据起始行，默认 header_row + 1
    - encoding: str | "auto" — CSV 编码，默认 "auto"

    当 parse_config 为 None 或空时，等价于 parse_file()。
    """
    if not parse_config or not isinstance(parse_config, dict):
        return parse_file(file_path)

    # 先用标准解析获取原始数据
    ext = Path(file_path).suffix.lower()

    if ext == ".csv":
        headers, all_rows = _parse_csv_all_rows(file_path, parse_config.get("encoding", "auto"))
    elif ext in (".xlsx", ".xls"):
        headers, all_rows = _parse_excel_all_rows(file_path)
    else:
        raise ValueError(f"不支持的文件格式: {ext}")

    # 应用 header_row / data_start_row 裁剪
    header_row = int(parse_config.get("header_row", 0))
    data_start = int(parse_config.get("data_start_row", header_row + 1))

    if header_row > 0 and header_row < len(all_rows):
        headers = [_clean_header(h) for h in all_rows[header_row]]
    elif header_row == 0:
        pass  # headers already from first row
    else:
        raise ValueError(f"header_row={header_row} 超出文件行数 {len(all_rows)}")

    # 截取数据行
    data_rows = all_rows[max(data_start, header_row + 1):]

    # 清理空行
    data_rows = [
        [cell if cell is not None else "" for cell in row]
        for row in data_rows
        if any(cell is not None and str(cell).strip() != "" for cell in row)
    ]

    return headers, data_rows


def _parse_csv_all_rows(file_path: str, encoding: str = "auto") -> tuple[list[str], list[list]]:
    """解析 CSV 返回全部行（不自动检测表头）"""
    encodings = ["utf-8", "utf-8-sig", "gbk", "gb2312", "gb18030", "latin-1"]
    content = None

    if encoding != "auto":
        encodings = [encoding] + [e for e in encodings if e != encoding]

    for enc in encodings:
        try:
            with open(file_path, "r", encoding=enc, newline="") as f:
                content = f.read()
            used_encoding = enc
            break
        except (UnicodeDecodeError, UnicodeError):
            continue

    if content is None:
        raise ValueError("无法识别 CSV 文件编码")

    lines = content.strip().splitlines()
    reader = csv.reader(lines)
    all_rows = list(reader)

    if len(all_rows) < 2:
        raise ValueError("CSV 文件至少需要表头行 + 一行数据")

    # 默认第一行是表头
    headers = [_clean_header(h) for h in all_rows[0]]
    return headers, all_rows


def _is_old_xls(file_path: str) -> bool:
    """通过文件头判断是否为旧 .xls 格式（OLE 或 BIFF）。"""
    try:
        with open(file_path, "rb") as f:
            header = f.read(8)
        # OLE2 复合文档 (D0 CF 11 E0 A1 B1 1A E1)
        if header[:4] == b'\xd0\xcf\x11\xe0':
            return True
        # BIFF 流 (09 08 ...)
        if header[:2] == b'\x09\x08':
            return True
        # BIFF 其他变体
        if header[:2] == b'\x09\x00' or header[:2] == b'\x09\x04':
            return True
        return False
    except Exception:
        return False


def _parse_excel_xlrd(file_path: str) -> tuple[list[str], list[list]]:
    """用 xlrd 解析旧 .xls 文件。"""
    import xlrd

    wb = xlrd.open_workbook(file_path, formatting_info=False)
    ws = wb.sheet_by_index(0)

    all_rows: list[list] = []
    for rx in range(ws.nrows):
        row = []
        for cx in range(ws.ncols):
            cell = ws.cell(rx, cx)
            if cell.ctype == xlrd.XL_CELL_EMPTY:
                row.append("")
            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                # 保留数值类型（整数或浮点数），方便后续金额解析
                val = cell.value
                if isinstance(val, float) and val == int(val):
                    row.append(int(val))
                else:
                    row.append(val)
            elif cell.ctype == xlrd.XL_CELL_DATE:
                row.append(str(cell.value))
            elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                row.append(bool(cell.value))
            elif cell.ctype == xlrd.XL_CELL_ERROR:
                row.append("")
            else:
                # 文本类型
                row.append(str(cell.value).strip())
        all_rows.append(row)

    if len(all_rows) < 2:
        raise ValueError("Excel .xls 文件至少需要表头行 + 一行数据")

    header_idx = _detect_header_row(all_rows)
    headers = [_clean_header(h) for h in all_rows[header_idx]]
    data_rows = all_rows[header_idx + 1:]

    # 清理：移除全空行
    data_rows = [
        [cell if cell is not None else "" for cell in row]
        for row in data_rows
        if any(cell is not None and str(cell).strip() != "" for cell in row)
    ]

    return headers, data_rows


def _parse_excel_all_rows(file_path: str) -> tuple[list[str], list[list]]:
    """解析 Excel 返回全部行。

    .xls 旧格式用 xlrd；.xlsx/.xlsm 用 openpyxl。
    openpyxl: read_only 模式对部分缺失默认样式的金蝶导出（Sheet1.max_row 异常返回 1）
    会失败，故优先用非 read_only 的 data_only 加载，异常时回退 read_only。
    """
    ext = Path(file_path).suffix.lower()

    # ── .xls 旧格式 → xlrd ──
    if ext == ".xls" and _is_old_xls(file_path):
        try:
            return _parse_xls_all_rows_xlrd(file_path)
        except Exception:
            # xlrd 失败时尝试自定义二进制解析（如 205201-2023.xls 的特殊格式）
            try:
                return _parse_xls_custom_binary(file_path)
            except Exception:
                raise  # 如果都失败则抛出原始 xlrd 错误

    # ── .xlsx/.xlsm → openpyxl ──
    from openpyxl import load_workbook

    all_rows: list[list] = []
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            all_rows.append(list(row))
        wb.close()
    except Exception:
        all_rows = []

    if len(all_rows) < 2:
        # 回退：用非 read_only data_only 模式重新读
        all_rows = []
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            all_rows.append(list(row))
        wb.close()

    if len(all_rows) < 2:
        raise ValueError("Excel 文件至少需要表头行 + 一行数据")

    headers = [_clean_header(h) for h in all_rows[0]]
    return headers, all_rows


def _parse_xls_all_rows_xlrd(file_path: str) -> tuple[list[str], list[list]]:
    """用 xlrd 解析旧 .xls 文件返回全部行（不做表头检测）。"""
    import xlrd

    wb = xlrd.open_workbook(file_path, formatting_info=False)
    ws = wb.sheet_by_index(0)

    all_rows: list[list] = []
    for rx in range(ws.nrows):
        row = []
        for cx in range(ws.ncols):
            cell = ws.cell(rx, cx)
            if cell.ctype == xlrd.XL_CELL_EMPTY:
                row.append("")
            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                val = cell.value
                if isinstance(val, float) and val == int(val):
                    row.append(int(val))
                else:
                    row.append(val)
            elif cell.ctype == xlrd.XL_CELL_DATE:
                row.append(str(cell.value))
            elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                row.append(bool(cell.value))
            elif cell.ctype == xlrd.XL_CELL_ERROR:
                row.append("")
            else:
                row.append(str(cell.value).strip())
        all_rows.append(row)

    if len(all_rows) < 2:
        raise ValueError("Excel .xls 文件至少需要表头行 + 一行数据")

    headers = [_clean_header(h) for h in all_rows[0]]
    return headers, all_rows


def _parse_xls_custom_binary(file_path: str) -> tuple[list[str], list[list]]:
    """解析自定义二进制格式（如 205201-2023.xls 的专有会计软件导出格式）。

    格式特点：字段元数据区和数据记录区均使用 04 02 标记的变长条目。
    本函数提取字段名后按变长条目读取数据记录。

    TASK-084 修复：这类文件同时包含 BIFF8 单元格记录（NUMBER/RK），
    其中 NUMBER 记录（0x0203）携带真实的金额数值，而 04 02 文本层
    只含科目代码/名称等文本字段、金额列为空。故在文本层解析完成后，
    扫描 BIFF8 NUMBER/RK 记录，按 (row, col) 把数值填入对应单元格。
    """
    with open(file_path, "rb") as f:
        data = f.read()

    if len(data) < 12:
        raise ValueError("文件太小，无法解析")

    pos = 12

    # ── 解析字段元数据 ──
    field_names: list[str] = []
    while pos + 12 <= len(data):
        marker = data[pos:pos+2]
        if marker == b'\x00\x00':
            pos += 2
            break
        if marker != b'\x04\x02':
            break
        total_size = int.from_bytes(data[pos+2:pos+4], 'little')
        name_len = int.from_bytes(data[pos+10:pos+12], 'little')
        name_bytes = data[pos+12:pos+12+name_len]
        try:
            name = name_bytes.decode('gbk').strip('\x00').strip()
        except Exception:
            name = ""
        field_names.append(name)
        pos += total_size + 4

    if not field_names:
        raise ValueError("无法识别文件字段结构")

    # 跳到数据区
    while pos < len(data) and data[pos] == 0:
        pos += 1

    # ── 读取数据记录 ──
    all_rows: list[list] = [field_names]
    while pos + 10 <= len(data):
        # 记录以 01 开头
        if data[pos] != 0x01:
            pos += 1
            continue
        record_start = pos
        pos += 1
        # 跳过记录头（9 字节）
        if pos + 9 > len(data):
            break
        pos += 9

        # 读取该记录的所有字段值
        row_values: dict[int, str] = {}
        while pos + 12 <= len(data):
            if data[pos:pos+2] == b'\x04\x02':
                total_size = int.from_bytes(data[pos+2:pos+4], 'little')
                field_idx = int.from_bytes(data[pos+6:pos+8], 'little')
                val_len = int.from_bytes(data[pos+10:pos+12], 'little')
                if pos + 12 + val_len > len(data):
                    break
                val_bytes = data[pos+12:pos+12+val_len]
                try:
                    val = val_bytes.decode('gbk').strip()
                except Exception:
                    val = val_bytes.decode('gbk', errors='replace').strip()
                row_values[field_idx] = val
                pos += total_size + 4
            elif data[pos:pos+2] == b'\x01\x02':
                # 下一个记录开始
                break
            else:
                pos += 1
                break

        # 按字段索引排列
        if row_values:
            row = [row_values.get(i, "") for i in range(len(field_names))]
            all_rows.append(row)
        else:
            pos = record_start + 1  # 跳过这个无效记录
            continue

    if len(all_rows) < 2:
        raise ValueError("未找到有效数据记录")

    # ── TASK-084：扫描 BIFF8 NUMBER/RK 记录，补充金额数值 ──
    # 文本层只含科目代码/名称，金额列（期初余额/借贷发生/期末结余）为空。
    # BIFF8 NUMBER 记录携带真实数值，按 (row, col) 填入 all_rows。
    _fill_biff_numeric_records(data, all_rows)

    headers = [_clean_header(h) for h in all_rows[0]]
    return headers, all_rows


def _fill_biff_numeric_records(data: bytes, all_rows: list[list]) -> None:
    """扫描 BIFF8 NUMBER(0x0203)/RK(0x027E)/MULRK(0x00BD) 记录，
    把数值按 (row, col) 填入 all_rows 对应单元格（仅当原值为空时）。

    BIFF 行号与 all_rows 索引一一对应：row 0 = 表头行，row 1+ = 数据行。
    """
    import struct

    pos = 0
    n = len(data)
    while pos + 4 <= n:
        rec_type = int.from_bytes(data[pos:pos+2], 'little')
        rec_len = int.from_bytes(data[pos+2:pos+4], 'little')
        body_start = pos + 4
        body_end = body_start + rec_len
        if body_end > n:
            break

        if rec_type == 0x0203 and rec_len >= 14:
            # NUMBER: row(2) + col(2) + xf(2) + value(8)
            row_idx = int.from_bytes(data[body_start:body_start+2], 'little')
            col_idx = int.from_bytes(data[body_start+2:body_start+4], 'little')
            val = struct.unpack('<d', data[body_start+6:body_start+14])[0]
            _set_numeric_cell(all_rows, row_idx, col_idx, val)

        elif rec_type == 0x027E and rec_len >= 10:
            # RK: row(2) + col(2) + xf(2) + rk(4)
            row_idx = int.from_bytes(data[body_start:body_start+2], 'little')
            col_idx = int.from_bytes(data[body_start+2:body_start+4], 'little')
            rk = int.from_bytes(data[body_start+6:body_start+10], 'little')
            val = _decode_rk(rk)
            _set_numeric_cell(all_rows, row_idx, col_idx, val)

        elif rec_type == 0x00BD and rec_len >= 6:
            # MULRK: row(2) + col_first(2) + (xf(2)+rk(4))*N + col_last(2)
            row_idx = int.from_bytes(data[body_start:body_start+2], 'little')
            col_first = int.from_bytes(data[body_start+2:body_start+4], 'little')
            col_last = int.from_bytes(data[body_end-2:body_end], 'little')
            rk_pos = body_start + 4
            for col_idx in range(col_first, col_last + 1):
                if rk_pos + 6 > body_end - 2:
                    break
                rk = int.from_bytes(data[rk_pos+2:rk_pos+6], 'little')
                val = _decode_rk(rk)
                _set_numeric_cell(all_rows, row_idx, col_idx, val)
                rk_pos += 6

        pos = body_end


def _decode_rk(rk: int) -> float:
    """解码 BIFF8 RK 值（4 字节压缩数值）。

    bit 0: 0=IEEE 754 double（高 30 位）, 1=signed integer（高 30 位）
    bit 1: 0=原值, 1=乘以 100
    """
    is_int = rk & 0x01
    is_scaled = (rk >> 1) & 0x01
    if is_int:
        # 高 30 位为有符号整数
        val = rk >> 2
        if val >= 0x20000000:  # 30 位有符号负数补码
            val -= 0x40000000
        value = float(val)
    else:
        # 高 30 位为 IEEE 754 double 的高 30 位，低 2 位补 0
        import struct
        packed = struct.pack('>I', rk) + b'\x00\x00\x00\x00'
        value = struct.unpack('>d', packed)[0]
    if is_scaled:
        value *= 100.0
    return value


def _set_numeric_cell(all_rows: list[list], row_idx: int, col_idx: int, val: float) -> None:
    """把数值填入 all_rows[row_idx][col_idx]，仅当原值为空/非数值时写入。"""
    if row_idx < 0 or row_idx >= len(all_rows):
        return
    row = all_rows[row_idx]
    while len(row) <= col_idx:
        row.append("")
    existing = row[col_idx]
    # 仅当原值为空或已是数值时写入，不覆盖文本层的文本值
    if existing is None or str(existing).strip() == "":
        row[col_idx] = val


def build_columns(headers: list[str], sample_rows: list[list] | None = None) -> list[dict]:
    """
    构建列描述符列表，每个列有稳定的 column_id。

    解决重复表头问题：column_id 基于列位置（col_001, col_002...），
    不受表头文本重复影响。duplicate_group 标记同名列的分组信息。

    Args:
        headers: 表头列表
        sample_rows: 前几行数据（用于生成 sample_values）

    Returns:
        [{"column_id": "col_001", "index": 0, "header": "...", ...}, ...]
    """
    # 第一遍：统计每个表头出现次数
    header_counts: dict[str, int] = {}
    for h in headers:
        key = (h or "").strip()
        header_counts[key] = header_counts.get(key, 0) + 1

    # 第二遍：构建列描述符
    occurrence_tracker: dict[str, int] = {}
    columns = []

    for i, header in enumerate(headers):
        normalized = header.strip() if header else ""
        col_id = f"col_{i + 1:03d}"

        col = {
            "column_id": col_id,
            "index": i,
            "header": header,
            "normalized_header": normalized,
            "sample_values": _extract_sample_values(sample_rows, i) if sample_rows else [],
        }

        # duplicate_group
        key = normalized
        total = header_counts.get(key, 1)
        if total > 1:
            occurrence_tracker[key] = occurrence_tracker.get(key, 0) + 1
            col["duplicate_group"] = {
                "header": normalized,
                "occurrence": occurrence_tracker[key],
                "total": total,
            }
        else:
            col["duplicate_group"] = None

        columns.append(col)

    return columns


def _extract_sample_values(sample_rows: list[list], col_index: int, max_samples: int = 3) -> list[str]:
    """从样例行中提取指定列的前几个值"""
    values = []
    for j in range(min(max_samples, len(sample_rows))):
        row = sample_rows[j]
        if col_index < len(row) and row[col_index] is not None:
            values.append(str(row[col_index]).strip())
        else:
            values.append("")
    return values
